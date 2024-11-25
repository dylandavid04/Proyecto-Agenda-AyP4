import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import re
import openpyxl

# Clases principales
class Contacto:
    def __init__(self, nombre, usuario, direccion, correo, telefono1, telefono2, fecha_nacimiento, grupos, password):
        self.nombre = nombre
        self.usuario = usuario
        self.direccion = direccion
        self.correo = correo
        self.telefono1 = telefono1
        self.telefono2 = telefono2
        self.fecha_nacimiento = fecha_nacimiento
        self.grupos = grupos
        self.password = password


class Nodo:
    def __init__(self, contacto):
        self.contacto = contacto
        self.siguiente = None


class ListaLigada:
    def __init__(self):
        self.cabeza = None

    def agregar(self, contacto):
        nuevo_nodo = Nodo(contacto)
        if not self.cabeza:
            self.cabeza = nuevo_nodo
        else:
            actual = self.cabeza
            while actual.siguiente:
                actual = actual.siguiente
            actual.siguiente = nuevo_nodo

    def eliminar(self, usuario, password):
        actual = self.cabeza
        anterior = None
        while actual:
            if actual.contacto.usuario == usuario and actual.contacto.password == password:
                if anterior:
                    anterior.siguiente = actual.siguiente
                else:
                    self.cabeza = actual.siguiente
                return True
            anterior = actual
            actual = actual.siguiente
        return False

    def buscar(self, usuario):
        actual = self.cabeza
        while actual:
            if actual.contacto.usuario == usuario:
                return actual.contacto
            actual = actual.siguiente
        return None

    def buscar_por_grupo(self, grupo):
        contactos = []
        actual = self.cabeza
        while actual:
            if grupo in actual.contacto.grupos:
                contactos.append(actual.contacto)
            actual = actual.siguiente
        return contactos

    def obtener_todos(self):
        contactos = []
        actual = self.cabeza
        while actual:
            contactos.append(actual.contacto)
            actual = actual.siguiente
        return contactos


# Validaciones
def validar_usuario(usuario):
    return re.fullmatch(r'[a-zA-Z][\w.-]{2,9}', usuario) is not None


def validar_password(password):
    return re.fullmatch(r'[A-Za-z\d@#$%^&+=!]{6,12}', password) is not None


def validar_direccion(direccion):
    return re.fullmatch(r'(Calle|Carrera|Avenida)\s+\d+.*', direccion) is not None

#^(Calle|Avenida Calle|Carrera|Avenida Carrera|Transversal|Avenida Transversal|Diagonal|Avenida Diagonal) 
# [1-9][0-9]{0,2}[A-Z]{0,1}( Bis( [A-Z]){0,1}){0,1}( Este| Sur){0,1}$

def validar_telefono(telefono):
    return re.fullmatch(r'\+57 \d{7,10}', telefono) is not None


def validar_correo(correo):
    return re.fullmatch(r'[\w.-]+@[\w.-]+\.\w+', correo) is not None


def validar_fecha(fecha):
    return re.fullmatch(r'(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[0-2])-\d{4}', fecha) is not None


# Funciones para leer y escribir en un archivo Excel
def cargar_contactos_desde_excel(agenda, archivo):
    try:
        wb = openpyxl.load_workbook(archivo)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nombre, usuario, password, direccion, correo, telefono1, telefono2, fecha_nacimiento, grupos = row
            grupos = set(grupos.split(","))
            contacto = Contacto(nombre, usuario, direccion, correo, telefono1, telefono2, fecha_nacimiento, grupos, password)
            agenda.agregar(contacto)
    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar el archivo Excel: {e}")


def guardar_contactos_a_excel(agenda, archivo):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Nombre", "Usuario", "Contraseña", "Dirección", "Correo", "Teléfono 1", "Teléfono 2", "Fecha de Nacimiento", "Grupos"])

        actual = agenda.cabeza
        while actual:
            contacto = actual.contacto
            grupos = ", ".join(contacto.grupos)
            sheet.append([contacto.nombre, contacto.usuario, contacto.password, contacto.direccion, contacto.correo,
                          contacto.telefono1, contacto.telefono2, contacto.fecha_nacimiento, grupos])
            actual = actual.siguiente

        wb.save(archivo)
    except Exception as e:
        messagebox.showerror("Error", f"Error al guardar el archivo Excel: {e}")


# Aplicación principal
class AgendaApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Agenda de Contactos")
        self.root.geometry("1000x600")
        self.agenda = ListaLigada()
        self.grupos = set(["Amigos", "Familia", "Trabajo"])
        self.archivo_excel = "contactos.xlsx"  # Archivo Excel donde se guardarán los contactos

        # Cargar contactos desde Excel al iniciar la aplicación
        cargar_contactos_desde_excel(self.agenda, self.archivo_excel)

        # Estilo visual
        self.estilo = ttk.Style()
        self.estilo.theme_use("clam")
        self.estilo.configure("Treeview", font=("Helvetica", 10))
        self.estilo.configure("Treeview.Heading", font=("Helvetica", 12, "bold"), background="#0052cc", foreground="white")

        # Encabezado
        tk.Label(root, text="Agenda de Contactos", font=("Helvetica", 18, "bold"), bg="#0052cc", fg="white").pack(fill=tk.X)

        # Tabla de contactos
        self.tree = ttk.Treeview(root, columns=("Nombre", "Usuario", "Grupos"), show="headings", height=15)
        self.tree.heading("Nombre", text="Nombre")
        self.tree.heading("Usuario", text="Usuario")
        self.tree.heading("Grupos", text="Grupos")
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Marco para los botones principales
        btn_frame = tk.Frame(root, bg="#f0f0f0")
        btn_frame.pack(fill=tk.X, pady=10)

        # Submarco para centrar los botones
        btn_inner_frame = tk.Frame(btn_frame, bg="#f0f0f0")
        btn_inner_frame.pack(anchor="center")

        # Botones principales (centrados)
        tk.Button(btn_inner_frame, text="Añadir Contacto", command=self.ventana_crear_contacto,
                  bg="#0052cc", fg="white", font=("Helvetica", 10)).grid(row=0, column=0, padx=10, pady=5)
        tk.Button(btn_inner_frame, text="Eliminar Contacto", command=self.eliminar_contacto,
                  bg="#cc0000", fg="white", font=("Helvetica", 10)).grid(row=0, column=1, padx=10, pady=5)
        tk.Button(btn_inner_frame, text="Editar Contacto", command=self.ventana_editar_contacto,
                  bg="#ff9900", fg="white", font=("Helvetica", 10)).grid(row=0, column=2, padx=10, pady=5)
        tk.Button(btn_inner_frame, text="Buscar Contacto", command=self.buscar_contacto,
                  bg="#006600", fg="white", font=("Helvetica", 10)).grid(row=0, column=3, padx=10, pady=5)
        tk.Button(btn_inner_frame, text="Buscar por Grupo", command=self.buscar_por_grupo,
                  bg="#663399", fg="white", font=("Helvetica", 10)).grid(row=0, column=4, padx=10, pady=5)
        tk.Button(btn_inner_frame, text="Importar desde Excel", command=self.importar_desde_excel,
                  bg="#0099cc", fg="white", font=("Helvetica", 10)).grid(row=0, column=5, padx=10, pady=5)

        # Actualizar tabla con los datos iniciales
        self.actualizar_tabla()

        # Guardar cambios al cerrar la aplicación
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    # Funciones principales de la aplicación
    def ventana_crear_contacto(self):
        ventana = tk.Toplevel(self.root)
        ventana.title("Añadir Contacto")
        ventana.geometry("400x500")

        tk.Label(ventana, text="Nombre").pack()
        entry_nombre = tk.Entry(ventana)
        entry_nombre.pack()

        tk.Label(ventana, text="Usuario").pack()
        entry_usuario = tk.Entry(ventana)
        entry_usuario.pack()

        tk.Label(ventana, text="Contraseña").pack()
        entry_password = tk.Entry(ventana, show="*")
        entry_password.pack()

        # Botón para mostrar/ocultar la contraseña
        def toggle_password():
            if entry_password.cget("show") == "*":
                entry_password.config(show="")
                btn_toggle_password.config(text="Ocultar")
            else:
                entry_password.config(show="*")
                btn_toggle_password.config(text="Ver")

        btn_toggle_password = tk.Button(ventana, text="Ver", command=toggle_password)
        btn_toggle_password.pack()

        tk.Label(ventana, text="Dirección").pack()
        entry_direccion = tk.Entry(ventana)
        entry_direccion.pack()

        tk.Label(ventana, text="Correo Electrónico").pack()
        entry_correo = tk.Entry(ventana)
        entry_correo.pack()

        tk.Label(ventana, text="Teléfono 1").pack()
        entry_telefono1 = tk.Entry(ventana)
        entry_telefono1.pack()

        tk.Label(ventana, text="Teléfono 2").pack()
        entry_telefono2 = tk.Entry(ventana)
        entry_telefono2.pack()

        tk.Label(ventana, text="Fecha de Nacimiento (DD-MM-AAAA)").pack()
        entry_fecha = tk.Entry(ventana)
        entry_fecha.pack()

        tk.Label(ventana, text="Grupos (coma separada)").pack()
        entry_grupos = tk.Entry(ventana)
        entry_grupos.pack()

        def agregar_contacto():
            nombre = entry_nombre.get()
            usuario = entry_usuario.get()
            password = entry_password.get()
            direccion = entry_direccion.get()
            correo = entry_correo.get()
            telefono1 = entry_telefono1.get()
            telefono2 = entry_telefono2.get()
            fecha = entry_fecha.get()
            grupos = set(entry_grupos.get().split(","))

            # Validaciones
            if not validar_usuario(usuario):
                messagebox.showerror("Error", "Usuario inválido.")
                return
            if not validar_password(password):
                messagebox.showerror("Error", "Contraseña inválida.")
                return
            if not validar_correo(correo):
                messagebox.showerror("Error", "Correo inválido.")
                return
            if not validar_telefono(telefono1) or (telefono2 and not validar_telefono(telefono2)):
                messagebox.showerror("Error", "Teléfonos inválidos.")
                return

            contacto = Contacto(nombre, usuario, direccion, correo, telefono1, telefono2, fecha, grupos, password)
            self.agenda.agregar(contacto)
            self.tree.insert("", tk.END, values=(nombre, usuario, ", ".join(grupos)))
            guardar_contactos_a_excel(self.agenda, self.archivo_excel)  # Guardar automáticamente en Excel
            messagebox.showinfo("Éxito", "Contacto añadido correctamente.")
            ventana.destroy()

        tk.Button(ventana, text="Agregar", command=agregar_contacto).pack()

    def eliminar_contacto(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            messagebox.showerror("Error", "Debe seleccionar un contacto.")
            return

        usuario = self.tree.item(seleccionado[0], "values")[1]
        password = simpledialog.askstring("Eliminar Contacto", "Ingrese la contraseña:", show="*")

        if not self.agenda.eliminar(usuario, password):
            messagebox.showerror("Error", "Contraseña incorrecta.")
        else:
            self.tree.delete(seleccionado[0])
            guardar_contactos_a_excel(self.agenda, self.archivo_excel)  # Guardar automáticamente en Excel
            messagebox.showinfo("Éxito", "Contacto eliminado correctamente.")

    def ventana_editar_contacto(self):
        seleccionado = self.tree.selection()
        if not seleccionado:
            messagebox.showerror("Error", "Debe seleccionar un contacto.")
            return

        usuario = self.tree.item(seleccionado[0], "values")[1]
        contacto = self.agenda.buscar(usuario)

        if not contacto:
            messagebox.showerror("Error", "Contacto no encontrado.")
            return

        password = simpledialog.askstring("Editar Contacto", "Ingrese la contraseña para editar:", show="*")

        if password != contacto.password:
            messagebox.showerror("Error", "Contraseña incorrecta.")
            return

        ventana = tk.Toplevel(self.root)
        ventana.title("Editar Contacto")
        ventana.geometry("400x500")

        tk.Label(ventana, text="Nombre").pack()
        entry_nombre = tk.Entry(ventana)
        entry_nombre.insert(0, contacto.nombre)
        entry_nombre.pack()

        tk.Label(ventana, text="Usuario").pack()
        entry_usuario = tk.Entry(ventana)
        entry_usuario.insert(0, contacto.usuario)
        entry_usuario.config(state=tk.DISABLED)
        entry_usuario.pack()

        tk.Label(ventana, text="Contraseña").pack()
        entry_password = tk.Entry(ventana, show="*")
        entry_password.insert(0, contacto.password)
        entry_password.pack()

        tk.Label(ventana, text="Dirección").pack()
        entry_direccion = tk.Entry(ventana)
        entry_direccion.insert(0, contacto.direccion)
        entry_direccion.pack()

        tk.Label(ventana, text="Correo Electrónico").pack()
        entry_correo = tk.Entry(ventana)
        entry_correo.insert(0, contacto.correo)
        entry_correo.pack()

        tk.Label(ventana, text="Teléfono 1").pack()
        entry_telefono1 = tk.Entry(ventana)
        entry_telefono1.insert(0, contacto.telefono1)
        entry_telefono1.pack()

        tk.Label(ventana, text="Teléfono 2").pack()
        entry_telefono2 = tk.Entry(ventana)
        entry_telefono2.insert(0, contacto.telefono2)
        entry_telefono2.pack()

        tk.Label(ventana, text="Fecha de Nacimiento (DD-MM-AAAA)").pack()
        entry_fecha = tk.Entry(ventana)
        entry_fecha.insert(0, contacto.fecha_nacimiento)
        entry_fecha.pack()

        tk.Label(ventana, text="Grupos (coma separada)").pack()
        entry_grupos = tk.Entry(ventana)
        entry_grupos.insert(0, ", ".join(contacto.grupos))
        entry_grupos.pack()

        def guardar_ediciones():
            nombre = entry_nombre.get()
            password = entry_password.get()
            direccion = entry_direccion.get()
            correo = entry_correo.get()
            telefono1 = entry_telefono1.get()
            telefono2 = entry_telefono2.get()
            fecha = entry_fecha.get()
            grupos = set(entry_grupos.get().split(","))

            # Validaciones
            if not validar_password(password):
                messagebox.showerror("Error", "Contraseña inválida.")
                return
            if not validar_telefono(telefono1) or (telefono2 and not validar_telefono(telefono2)):
                messagebox.showerror("Error", "Teléfonos inválidos.")
                return

            contacto.nombre = nombre
            contacto.password = password
            contacto.direccion = direccion
            contacto.correo = correo
            contacto.telefono1 = telefono1
            contacto.telefono2 = telefono2
            contacto.fecha_nacimiento = fecha
            contacto.grupos = grupos

            self.tree.item(seleccionado[0], values=(nombre, contacto.usuario, ", ".join(grupos)))
            messagebox.showinfo("Éxito", "Contacto actualizado correctamente.")
            ventana.destroy()

        tk.Button(ventana, text="Guardar", command=guardar_ediciones).pack()

    def buscar_contacto(self):
        usuario = simpledialog.askstring("Buscar Contacto", "Ingrese el usuario a buscar:")
        contacto = self.agenda.buscar(usuario)

        if contacto:
            messagebox.showinfo("Contacto Encontrado", f"Nombre: {contacto.nombre}\nCorreo: {contacto.correo}\nTeléfonos: {contacto.telefono1}, {contacto.telefono2}\nFecha de Nacimiento: {contacto.fecha_nacimiento}\nGrupos: {', '.join(contacto.grupos)}")
        else:
            messagebox.showerror("Error", "Contacto no encontrado.")

    def buscar_por_grupo(self):
        grupo = simpledialog.askstring("Buscar por Grupo", "Ingrese el nombre del grupo:")
        contactos = self.agenda.buscar_por_grupo(grupo)

        if contactos:
            mensaje = "\n".join([f"{contacto.nombre} - {contacto.usuario}" for contacto in contactos])
            messagebox.showinfo("Contactos en el Grupo", mensaje)
        else:
            messagebox.showerror("Error", "No se encontraron contactos en este grupo.")

    
    def importar_desde_excel(self):
        archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
        if archivo:
            cargar_contactos_desde_excel(self.agenda, archivo)
            self.actualizar_tabla()
            messagebox.showinfo("Éxito", "Contactos importados desde Excel correctamente.")

    def actualizar_tabla(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        contactos = self.agenda.obtener_todos()
        for contacto in contactos:
            grupos = ", ".join(contacto.grupos)
            self.tree.insert("", "end", values=(contacto.nombre, contacto.usuario, grupos))

    def on_closing(self):
        guardar_contactos_a_excel(self.agenda, self.archivo_excel)
        self.root.quit()

root = tk.Tk()
app = AgendaApp(root)
root.mainloop()
